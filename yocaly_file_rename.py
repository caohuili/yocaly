import os,pprint,sys
sys.path.append('..\\..')
import openpyxl,xlrd


def get_patient_name(date):
	#yocaly_path = os.getcwd()+'\\yocaly_download\\'
	yocaly_path = '..\\..\\yocaly_download\\'
	path = yocaly_path+date+'\\'
	udate = date[:4]+'-'+date[4:6]+'-'+date[6:]
	filenames = os.listdir(path)
	#path = 'E:\lepu_yocaly_xmlpdf\yocaly_zip\20170817'
	excel_xfile = path+udate+'.xlsx'
	excel_file = path+udate+'.xls'
	#print(excel_path)
	new_name = {}
	if udate+'.xlsx' in filenames:
		wb = openpyxl.load_workbook(excel_xfile)
		sheet = wb.get_sheet_by_name('第一页')

		lines = sheet.max_row
		
		for line in range(2,lines+1):
			name = sheet.cell(row=line,column=1).value
			patient_id = sheet.cell(row=line,column=2).value
			print('yocaly文件xml/pdf %s重命名'%name)
			new_name[patient_id] = name.replace(' ','').replace(',','').replace('.','')

		print(new_name)
	elif udate+'.xls' in filenames:
		wb = xlrd.open_workbook(excel_file)
		sheet = wb.sheet_by_name('第一页')
		lines = sheet.nrows
		for line in range(1,lines):
			name = sheet.cell(line,0).value
			patient_id = sheet.cell(line,1).value
			patient_state = sheet.cell(line,2).value
			if '不做处理' in patient_state or '71' in patient_state:
				pass
			else:
				print('yocaly文件xml/pdf %s重命名'%name)
				new_name[patient_id] = name.replace(' ','').replace(',','').replace('.','')
		print(new_name)

	with open ('D:\\ScriptData\\test_patient.txt','w',encoding='gbk') as f:
		for patient_id,name in new_name.items():
			print(name+','+patient_id,file=f)
	#new_name_path = path[:-2]
	with open(('run_result_file\\patientid_dict.py'),'w',encoding='utf-8')as df:
		df.write('patientid_dict='+pprint.pformat(new_name))


		'''
		new_name = {'56B6A2F7F7B6897850128354EAB08F0B':'林文清',
		'A2B5BD26E6D5CF9579F800A703E7DA49':'洪启冲',
		'59BE45FE78F5E2E87B11AC4EC84C81B9':'孙成灿',
		'0F9FF7E1A6EBB8FCF1C039B74B0DA21D':'黄仙丽',
		'42DB5EC8341E801D87DF912524A6E9BB':'陈强明',
		'3F458644C434C4F5F2DE1109922B230F':'余乃森',
		'0C9BC65D20EED285C212ED5EDFC6A1FF':'张国坤',
		'D41B0B7FE7C5B9FB10B6274D8303E5FF':'林金宝'}
	
	'''
	for filename in filenames:
		name = filename.split('.')[0][:-1]
		if '.DAT.XML' in filename:
			try:
				os.rename(path+filename,path+new_name[name]+'.XML')
			except:
				pass

		if '.DAT.PDF' in filename:
			try:
				os.rename(path+filename,path+new_name[name]+'.PDF')
			except:
				pass

	return new_name
if __name__ == '__main__':
	#path = 'E:\\lepu_yocaly_xmlpdf\\yocaly_zip\\'
	#lepu_path = 'D:\\ScriptData\\datas\\'
	get_patient_name('20170925')


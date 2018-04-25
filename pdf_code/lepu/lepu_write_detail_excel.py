import openpyxl,re
from run_result_file import lepu_detail_dict


def lepu_write_detail_excel2(before_excel_file, result_excel, date):
	# users = ['蔡新民_全流程', '陈朝元_全流程', '陈冬和_全流程', '丁小婷_全流程', '冯殿玉_全流程', '龚道玲_全流程', '龚道玲2_全流程', '侯秀芳_全流程', '胡才刚_全流程', '金友福_全流程', '李会敏_全流程', '李建美_全流程', '李丽琼_全流程', '李纳_全流程', '李先林_全流程', '李懿庄_全流程', '刘桂花_全流程', '刘余生_全流程', '刘振华_全流程', '罗洛_全流程', '马建民_全流程', '潘永泉_全流程', '邱凤_全流程', '全玉书_全流程', '唐世兰_全流程', '王银妹_全流程', '魏文博_全流程', '谢成阶_全流程', '熊美秀_全流程', '徐福喜_全流程', '周林新_全流程', '朱付青_全流程', '朱家东_全流程', '朱守禄_全流程', '朱秀英_全流程']

	excel_tag = ['基本心律', '早', '成对', '二联律', '三联律', '心动过速', '加速', '室上速', '交界性逸搏', '房性逸搏', '心房扑动-心房颤动', '早', '成对', '二联律',
				 '三联律', '心动过速', '加速', '室性逸搏', '心室预激', '一度房室', '二度Ⅰ型房室', '二度Ⅱ型房室', '三度房室', '二度Ⅰ型窦房', '二度Ⅱ型窦房', '右束支', '左束支',
				 '室内阻滞', '平均心率(bpm)', '最大心率(bpm)', '最大心率(bpm)发生时间', '最小心率(bpm)', '最小心率(bpm)发生时间', '总心搏数',
				 '房扑-房颤(占总心搏)%', '最长RR间期(s)', '最长RR间期(s)发生时间', '室上性总数', '室上早数', '成对室上早数', '二联律数', '三联律数', '房性逸搏数',
				 '交界性逸搏数', '室性总数', '室早数', '成对室早数', '二联律数', '三联律数', '室性逸搏数']
	conclusion_tags = ['室上性心律失常', '室性心律失常', '传导阻滞','心室预激' ]
	conclusion_tag1 = ['早', '成对', '二联律', '三联律', '心动过速', '加速', '室上速', '交界性逸搏', '房性逸搏', '心房扑动-心房颤动']
	conclusion_tag2 = ['早', '成对', '二联律', '三联律', '心动过速', '加速', '室性逸搏']
	conclusion_tag3 = ['一度房室', '二度Ⅰ型房室', '二度Ⅱ型房室', '三度房室', '二度Ⅰ型窦房', '二度Ⅱ型窦房', '完全性右束支', '完全性左束支', '室内阻滞']

	total_tag = ['平均心率(bpm)', '最大心率(bpm)', '最大心率发生时间', '最小心率(bpm)', '最小心率发生时间', '总心搏数', '房扑-房颤(占总心搏)%', '最长RR间期(s)',
				 '最长RR间期发生时间', '室上性总数', '室上早数', '成对室上早数', '二联律数', '三联律数', '房性逸搏数', '交界性逸搏数', '室性总数', '室早数', '成对室早数',
				 '二联律数', '三联律数', '室性逸搏数']

	detail_dict = lepu_detail_dict.lepu_detail_dict
	wb = openpyxl.load_workbook(before_excel_file)
	sheet = wb.get_sheet_by_name('pdf-result')

	lines = sheet.max_row
	total_tag_num = len(total_tag)
	line=2
	for name,details in detail_dict.items():
		sheet.cell(row=line + 1, column=1).value = date
		sheet.cell(row=line + 1, column=2).value = name.replace('全流程', 'AI')
		content = detail_dict[name]
		conclusion = content[-1]['结论']
		basic_rhythm = conclusion[3]
		basic_rhythm = basic_rhythm.replace(':', '').replace('：', '')
		sheet.cell(row=line + 1, column=3).value = basic_rhythm

		# 写入各项统计数值
		for i in range(total_tag_num):
			sheet.cell(row=line + 1, column=i + 31).value = content[i][total_tag[i]]

		if conclusion_tags[0] in conclusion:
			pos1 = conclusion.index(conclusion_tags[0])
			text1 = conclusion[pos1 + 1]
			print(text1)
			for i in range(1, 11):
				if excel_tag[i] in text1:
					sheet.cell(row=line + 1, column=i + 3).value = 'Y'
				else:
					pass
		else:
			pass
		if conclusion_tags[1] in conclusion:
			pos2 = conclusion.index(conclusion_tags[1])
			text2 = conclusion[pos2 + 1]
			print(text2)
			for i in range(11, 18):
				if excel_tag[i] in text2:
					sheet.cell(row=line + 1, column=i + 3).value = 'Y'
				else:
					pass
		else:
			pass
		if conclusion_tags[2] in conclusion:
			pos3 = conclusion.index(conclusion_tags[2])
			text3 = conclusion[pos3 + 1]
			r_f, l_f = judge_rl_incomplete(text3)
			for i in range(19, 28):
				if excel_tag[i] in text3:
					sheet.cell(row=line + 1, column=i + 3).value = 'Y'
					if i == 25 and r_f == 1:
						sheet.cell(row=line + 1, column=i + 3).value = ''
					else:
						pass
					if i == 26 and l_f == 1:
						sheet.cell(row=line + 1, column=i + 3).value = ''
					else:
						pass

				else:
					pass
		else:
			pass

		if conclusion_tags[3] in conclusion:
			sheet.cell(row=line + 1, column=21).value = 'Y'
		else:
			pass

		line+=1

	wb.save(result_excel)

def lepu_write_detail_excel(before_excel_file,result_excel,date):
	#users = ['蔡新民_全流程', '陈朝元_全流程', '陈冬和_全流程', '丁小婷_全流程', '冯殿玉_全流程', '龚道玲_全流程', '龚道玲2_全流程', '侯秀芳_全流程', '胡才刚_全流程', '金友福_全流程', '李会敏_全流程', '李建美_全流程', '李丽琼_全流程', '李纳_全流程', '李先林_全流程', '李懿庄_全流程', '刘桂花_全流程', '刘余生_全流程', '刘振华_全流程', '罗洛_全流程', '马建民_全流程', '潘永泉_全流程', '邱凤_全流程', '全玉书_全流程', '唐世兰_全流程', '王银妹_全流程', '魏文博_全流程', '谢成阶_全流程', '熊美秀_全流程', '徐福喜_全流程', '周林新_全流程', '朱付青_全流程', '朱家东_全流程', '朱守禄_全流程', '朱秀英_全流程']
	
	excel_tag = ['基本心律','早','成对','二联律','三联律','心动过速','加速','室上速','交界性逸搏','房性逸搏','心房扑动-心房颤动','早','成对','二联律','三联律','心动过速','加速','室性逸搏','心室预激','一度房室','二度Ⅰ型房室','二度Ⅱ型房室','三度房室','二度Ⅰ型窦房','二度Ⅱ型窦房','右束支','左束支','室内阻滞','平均心率(bpm)','最大心率(bpm)','最大心率(bpm)发生时间','最小心率(bpm)','最小心率(bpm)发生时间','总心搏数','房扑-房颤(占总心搏)%','最长RR间期(s)','最长RR间期(s)发生时间','室上性总数','室上早数','成对室上早数','二联律数','三联律数','房性逸搏数','交界性逸搏数','室性总数','室早数','成对室早数','二联律数','三联律数','室性逸搏数']
	conclusion_tags = ['室上性心律失常','室性心律失常','传导阻滞','心室预激']
	conclusion_tag1 = ['早','成对','二联律','三联律','心动过速','加速','室上速','交界性逸搏','房性逸搏','心房扑动-心房颤动']
	conclusion_tag2 = ['早','成对','二联律','三联律','心动过速','加速','室性逸搏']
	conclusion_tag3 = ['一度房室','二度Ⅰ型房室','二度Ⅱ型房室','三度房室','二度Ⅰ型窦房','二度Ⅱ型窦房','完全性右束支','完全性左束支','室内阻滞']

	total_tag = ['平均心率(bpm)','最大心率(bpm)','最大心率发生时间','最小心率(bpm)','最小心率发生时间','总心搏数','房扑-房颤(占总心搏)%','最长RR间期(s)','最长RR间期发生时间','室上性总数','室上早数','成对室上早数','二联律数','三联律数','房性逸搏数','交界性逸搏数','室性总数','室早数','成对室早数','二联律数','三联律数','室性逸搏数']

	detail_dict = lepu_detail_dict.lepu_detail_dict
	wb = openpyxl.load_workbook(before_excel_file)
	sheet = wb.get_sheet_by_name('pdf-result')


	lines = sheet.max_row
	total_tag_num = len(total_tag)

	for line in range(3,lines+1,2):
		name = sheet.cell(row=line,column=2).value
		#name = name + '_全流程'

		if name in detail_dict:
			sheet.cell(row=line+1,column=1).value = date
			#sheet.cell(row=line+1,column=2).value = name.replace('全流程', 'AI')
			sheet.cell(row=line + 1, column=2).value = name+'_AI'
			content = detail_dict[name]
			conclusion = content[-1]['结论']
			basic_rhythm = conclusion[3]
			basic_rhythm=basic_rhythm.replace(':','').replace('：','')
			sheet.cell(row=line+1, column=3).value = basic_rhythm


			#写入各项统计数值
			for i in range(total_tag_num):
				sheet.cell(row=line+1,column=i+31).value = content[i][total_tag[i]]

			if conclusion_tags[0] in conclusion:
				pos1 = conclusion.index(conclusion_tags[0])
				text1 = conclusion[pos1+1]
				print(text1)
				for i in range(1,11):
					if excel_tag[i] in text1:
						sheet.cell(row=line+1,column=i+3).value = 'Y'
					else:
						pass
			else:
				pass
			if conclusion_tags[1] in conclusion:
				pos2 = conclusion.index(conclusion_tags[1])
				text2 = conclusion[pos2 + 1]
				print(text2)
				for i in range(11,18):
					if excel_tag[i] in text2:
						sheet.cell(row=line+1,column=i+3).value = 'Y'
					else:
						pass
			else:
				pass
			if conclusion_tags[2] in conclusion:
				pos3 = conclusion.index(conclusion_tags[2])
				text3 = conclusion[pos3 + 1]
				r_f,l_f = judge_rl_incomplete(text3)
				for i in range(19, 28):
					if excel_tag[i] in text3:
						sheet.cell(row=line+1, column=i + 3).value = 'Y'
						if i==25 and r_f==1:
							sheet.cell(row=line+1, column=i + 3).value = ''
						else:
							pass
						if i==26 and l_f==1:
							sheet.cell(row=line+1, column=i + 3).value = ''
						else:
							pass

					else:
						pass
			else:
				pass

			if conclusion_tags[3] in conclusion:
				sheet.cell(row=line + 1, column=21).value = 'Y'
			else:
				pass
		else:
			pass

	wb.save(result_excel)


def judge_rl_incomplete(text):
	r_s1='完全性?右束支'
	r_s2='非完全性?右束支'
	l_s1='完全性?左束支'
	l_s2='非完全性?左束支'
	r1=re.findall(r'%s'%r_s1, text)
	l1=re.findall(r'%s'%l_s1, text)
	r2=re.findall(r'%s'%r_s2, text)
	l2=re.findall(r'%s'%l_s2, text)
	if len(r1)==1 and len(r2)==1:
		r_f = 1
	else:
		r_f=0
	if len(l1)==1 and len(l2)==1:
		l_f = 1
	else:
		l_f=0
	return r_f,l_f



if __name__ == '__main__':
	date = '20170926'
	excel_path = '..\\..\\run_result_file\\'
	before_excel_file = excel_path + 'yocaly_pdf_detail.xlsx'
	result_excel = excel_path + date + '_'+'pdf_result.xlsx'
	lepu_write_detail_excel(before_excel_file,result_excel,date)

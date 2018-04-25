import openpyxl
from run_result_file import yocaly_detail_dict


def yocaly_write_detail_excel(modle_excel,file_excel):
    excel_tag = ['基本心律','室上性早搏','成对','二联律','三联律','室上性心动过速','加速性室上性','室上速','交界性逸搏','室上性逸搏','心房扑动-颤动','室性早搏','成对','二联律','三联律','室性心动过速','加速性室性','室性逸搏','心室预激','一度房室','二度Ⅰ型房室','二度Ⅱ型房室','三度房室','二度Ⅰ型窦房','二度Ⅱ型窦房','完全性右束支','完全性左束支','室内阻滞']
    three_tag = ['早','成对','二联律','三联律']

    tag = ['平均心率(bpm)','最大心率(bpm)','最大心率发生时间','最小心率(bpm)','最小心率发生时间','总心搏数','房扑/房颤占时比','最长停搏','最长RR间期发生时间','室上性总数','室上早','成对室上早','二联律','三联律','房性逸搏','交界性逸搏','室性总数','室早','成对室早','二联律','三联律','室性逸搏']
    total_tag_num = len(tag)

    yocaly_dict = yocaly_detail_dict.yocaly_detail_dict
    #yocaly_dict = conclusion_dict.conclusion_dict
    wb = openpyxl.load_workbook(modle_excel)



    sheet = wb.get_sheet_by_name('pdf-result')

    lines = len(yocaly_dict)
    tag_list = list(range(1,30))
    tag_num = tag_list[6:12]+tag_list[16:]#[7, 8, 9, 10, 11, 12, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28,29]


    line = 3

    keys = yocaly_dict.keys()
    for id,text in yocaly_dict.items():
        #content = yocaly_dict[name]
        sheet.cell(row=line,column=2).value = id
        content = text[-2]['结论']
        sheet.cell(row=line,column=3).value = content[0]

        text_num = len(content)

        for i in range(total_tag_num):
            sheet.cell(row=line,column=i+31).value = text[i][tag[i]]

        for n in range(text_num):
            for m in tag_num:
                if excel_tag[m-2] in content[n]:
                    sheet.cell(row=line, column=m+1).value = 'Y'
                else:
                    pass
            if  '颤动' in content[n] or '扑动' in content[n]:
                sheet.cell(row=line, column=13).value = 'Y'
            else:
                pass

            if '二度房室' in content[n]:
                #sheet.merge('w%d:x%d'%(line,line))
                sheet.cell(row=line, column=23).value = '未分子型'
                sheet.cell(row=line, column=24).value = '未分子型'

            if '二度窦房' in content[n]:
                #sheet.merge('w%d:x%d'%(line,line))
                sheet.cell(row=line, column=26).value = '未分子型'
                sheet.cell(row=line, column=27).value = '未分子型'

            if '室性' in content[n]:
                for reput_j in range(13, 17):
                    if three_tag[reput_j - 13] in content[n]:
                        sheet.cell(row=line, column=reput_j+1).value = 'Y'
                    else:
                        pass


            elif ('室性' not in content[n]) and ('室上性' in content[n]):
                for reput1_j in range(3, 7):
                    if three_tag[reput1_j - 3] in content[n]:
                        sheet.cell(row=line, column=reput1_j+1).value = 'Y'
                    else:
                        pass
            else:
                pass

        line += 2

    wb.save(file_excel)
if __name__ == '__main__':

    modle_excel = '..\\pdf_result.xlsx'
    file_excel = '..\\..\\run_result_file\\yocaly_pdf_detail.xlsx'
    yocaly_write_detail_excel(modle_excel,file_excel)
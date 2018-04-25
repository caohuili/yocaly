import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

def get_excel_result(d):
    labels=['N', 'S', 'V', 'N_CLB', 'N_CRB', 'SE', 'JE', 'VE', 'P', 'Q', 'O', 'X']
    wb = openpyxl.Workbook()
    #sheet = wb.create_sheet('result')
    sheet = wb.get_active_sheet()
    line=2


    for row_label in labels:
        col = 2
        sheet.cell(row=line,column=1).value = row_label
        sheet.cell(row=1, column=line).value = row_label.lower()
        for col_label in labels:
            col_label = col_label.lower()
            excel_label = row_label + '-' + col_label
            if d[excel_label]==0:
                pass
            else:
                sheet.cell(row=line, column=col).value = d[excel_label]
            col+=1
        line+=1

    sheet.cell(row=15, column=2).value = '敏感率'
    sheet.cell(row=15, column=3).value = '阳性率'
    sheet.cell(row=16, column=1).value = 'QRS'
    sheet.cell(row=16, column=2).value = '=SUM(B2:K11)/SUM(B2:M11)'
    sheet.cell(row=16, column=3).value = '=SUM(B2:K11)/SUM(B2:K13)'

    for i in range(0,len(labels)-3):
        sheet.cell(row=17+i, column=1).value = labels[i]
        # sheet.cell(row=17 + i, column=2).value = '=B2/SUM(B2:M2)'
        # sheet.cell(row=17 + i, column=3).value = '=B2/SUM(B2:B13)'
        col_num = get_column_letter(i+2)
        sen_label = labels[i]+'-'+labels[i].lower()
        if d[sen_label] != 0:
            sheet.cell(row=17 + i, column=2).value = '=%s%d/SUM(B%d:M%d)'%(col_num,i+2,i+2,i+2)
            sheet.cell(row=17 + i, column=3).value = '=%s%d/SUM(%s2:%s13)'%(col_num,i+2,col_num,col_num)
        else:
            pass

    wb.save('result.xlsx')

if __name__ == '__main__':
    d={'X-v': 0, 'S-v': 166, 'P-p': 0, 'V-n': 345, 'V-ve': 39, 'X-q': 0, 'N_CLB-x': 0, 'N_CRB-p': 0, 'Q-n_crb': 0, 'SE-ve': 0, 'X-je': 0, 'P-q': 1285, 'O-ve': 7, 'N_CLB-n': 296, 'X-n_crb': 0, 'P-s': 31, 'S-q': 269, 'Q-q': 4, 'JE-q': 5, 'Q-p': 0, 'V-je': 0, 'VE-x': 0, 'S-x': 0, 'N_CLB-v': 808, 'N-p': 0, 'N_CLB-je': 0, 'N_CLB-n_clb': 73, 'O-v': 601, 'X-s': 0, 'X-n': 0, 'SE-je': 0, 'JE-o': 0, 'N_CRB-n': 290, 'N_CLB-o': 8, 'O-s': 2109, 'VE-p': 0, 'N_CLB-ve': 78, 'X-x': 0, 'VE-q': 0, 'P-je': 0, 'N-n_crb': 44, 'O-n': 215, 'JE-se': 0, 'VE-o': 0, 'N-o': 396, 'P-x': 0, 'V-n_clb': 1, 'N-se': 0, 'N_CRB-ve': 10, 'S-n': 117, 'SE-o': 0, 'N_CLB-s': 2, 'Q-n': 4, 'Q-ve': 1, 'JE-n_clb': 1, 'P-n': 457, 'X-ve': 0, 'O-o': 0, 'V-s': 42, 'JE-s': 34, 'N_CRB-o': 2, 'SE-n_crb': 0, 'JE-ve': 2, 'VE-ve': 50, 'X-n_clb': 0, 'O-je': 0, 'VE-s': 0, 'S-p': 0, 'V-q': 195, 'O-n_clb': 18, 'V-n_crb': 10, 'N_CRB-s': 11, 'S-s': 1200, 'P-se': 0, 'Q-s': 1, 'X-p': 0, 'N_CRB-v': 2, 'Q-o': 0, 'SE-n_clb': 0, 'Q-x': 0, 'S-o': 68, 'N_CLB-p': 0, 'O-q': 341, 'N_CLB-q': 227, 'JE-p': 0, 'N-n_clb': 975, 'VE-je': 0, 'N-je': 0, 'Q-v': 4, 'SE-n': 0, 'N-x': 0, 'O-x': 0, 'N-v': 108, 'N_CRB-n_crb': 6667, 'X-o': 0, 'S-n_crb': 954, 'P-ve': 0, 'Q-n_clb': 0, 'JE-n_crb': 0, 'SE-x': 0, 'VE-n_clb': 1, 'V-se': 0, 'JE-n': 183, 'VE-v': 55, 'O-n_crb': 5, 'SE-se': 0, 'V-x': 0, 'VE-n': 0, 'N_CRB-se': 0, 'JE-v': 4, 'X-se': 0, 'S-je': 0, 'Q-je': 0, 'N_CRB-n_clb': 270, 'S-se': 0, 'P-n_clb': 39, 'JE-je': 0, 'O-se': 0, 'N_CLB-se': 0, 'P-n_crb': 57, 'SE-s': 0, 'N_CRB-je': 0, 'N_CRB-q': 0, 'O-p': 0, 'N-q': 12284, 'N_CRB-x': 0, 'P-o': 1, 'S-ve': 3, 'VE-se': 0, 'SE-q': 0, 'V-v': 7166, 'Q-se': 0, 'SE-v': 0, 'N-ve': 4, 'V-p': 0, 'N-s': 702, 'VE-n_crb': 0, 'JE-x': 0, 'SE-p': 0, 'V-o': 71, 'P-v': 294, 'N-n': 57510, 'S-n_clb': 3, 'N_CLB-n_crb': 403}
    get_excel_result(d)
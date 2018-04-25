import openpyxl,os,pprint
import math

from run_result_file import yocaly_r_dict
from run_result_file import lepu_r_dict


def write_r_excel(date):
	wb = openpyxl.load_workbook(os.getcwd()+'\\xml_code\\rpos_label.xlsx')
	yl_diff_dict,move_list = get_r_result()
	excel_tag = ['错检r_pos','yocaly','lepu','漏检r_pos','label','多检r_pos','label']
	for key,value in yl_diff_dict.items():
		sheet = wb.create_sheet(key,0)
		for i in range(len(excel_tag)):
			sheet.cell(row=1,column=i+1).value = excel_tag[i]
		errorline = 2
		for error in value[0]:

			sheet.cell(row=errorline,column=1).value = error[0]
			sheet.cell(row=errorline,column=2).value = error[1]
			sheet.cell(row=errorline,column=3).value = error[2]
			errorline += 1
		rows = len(value[0])

		lessline = 2
		for less in value[1]:
			sheet.cell(row=lessline,column=4).value = less[0]
			sheet.cell(row=lessline,column=5).value = less[1]
			lessline += 1

		moreline = 2
		for more in value[2]:
			sheet.cell(row=moreline,column=6).value = more[0]
			sheet.cell(row=moreline,column=7).value = more[1]
			moreline += 1

		if bool(move_list):
			for move_p in move_list:
				sheet.cell(row=rows+2,column=1).value = move_p[2]
				sheet.cell(row=rows+2,column=2).value = move_p[1]
				sheet.cell(row=rows+2,column=3).value = move_p[3]
				rows+=1


	wb.save(os.getcwd()+'\\run_result_file\\'+date+'_xml_result.xlsx')





def get_r_result():
	
	yocaly_rpos_dict = yocaly_r_dict.ecg_r_dict
	lepu_rpos_dict = lepu_r_dict.ecg_r_dict
	yl_diff_dict = {}
	for y_key,y_value in yocaly_rpos_dict.items():
		print(y_key)

		if y_key in lepu_rpos_dict:
			l_value = lepu_rpos_dict[y_key]
			
			yuser_r_dict ={}
			luser_r_dict ={}

			for i in y_value:
				yuser_r_dict.update(i)

			for j in l_value:
				luser_r_dict.update(j)
			yuser_r_set = set(yuser_r_dict)
			luser_r_set = set(luser_r_dict)

			yluser_r_set = yuser_r_set&luser_r_set
			yluser_r_list = list(yluser_r_set)
			yluser_r_list.sort()

			more_r_set=luser_r_set-yluser_r_set
			less_r_set=yuser_r_set-yluser_r_set
			more_r_list=list(more_r_set)
			less_r_list=list(less_r_set)
			more_and_less_list = more_r_list+less_r_list
			more_and_less_list.sort()
			re_more_r_list=[]
			re_less_r_list=[]
			move_r_list=[]
			re_move_r_list=[]
			for i in range(len(more_and_less_list)-1):
				if (more_and_less_list[i+1]-more_and_less_list[i])<21:
					re_move_r_list+=[more_and_less_list[i],more_and_less_list[i+1]]
					if more_and_less_list[i] in more_r_list:
						if luser_r_dict[more_and_less_list[i]]==yuser_r_dict[more_and_less_list[i+1]]:
							pass
						else:
							move_r_list.append([more_and_less_list[i],more_and_less_list[i+1]])
					else:
						if luser_r_dict[more_and_less_list[i+1]]==yuser_r_dict[more_and_less_list[i]]:
							pass
						else:
							move_r_list.append([more_and_less_list[i+1],more_and_less_list[i]])

				else:
					if more_and_less_list[i] in more_r_list:
						re_more_r_list.append(more_and_less_list[i])
					else:
						re_less_r_list.append(more_and_less_list[i])
			re_more_r_list=list(set(re_more_r_list)-set(re_move_r_list))
			re_less_r_list=list(set(re_less_r_list)-set(re_move_r_list))


			re_more_r_list.sort()

			re_less_r_list.sort()
			move_r_list.sort()

			more_label_list = []
			less_label_list = []
			error_label_list = []
			move_label_list = []
			for mr in re_more_r_list:
				more_label_list.append([mr,luser_r_dict[mr]])
			for lr in re_less_r_list:
				less_label_list.append([lr,yuser_r_dict[lr]])

			if bool(move_r_list):
				for move_r in move_r_list:
					move_label_list.append([move_r[1],yuser_r_dict[move_r[1]],move_r[0],luser_r_dict[move_r[0]]])
			else:pass


			for ylr in yluser_r_list:
				if luser_r_dict[ylr] != yuser_r_dict[ylr]:
					error_label_list.append([ylr,yuser_r_dict[ylr],luser_r_dict[ylr]])
				else:
					pass
			yl_diff_dict[y_key]=[error_label_list,less_label_list,more_label_list]
			#print(error_r_dict)


		else:
			print('乐普数据无此人')

	return yl_diff_dict,move_label_list


def get_lot_error_dict(num):
	diff_dict,move_list = get_r_result()
	#diff_list = diff_dict[fileid]
	all_draw_pos ={}
	for name,diff_list in diff_dict.items():
		print(name)
		draw_pos={}
		for i in range(len(diff_list[0])):
			flag = diff_list[0][i][1]+'--'+diff_list[0][i][2]
			#new_flag = diff_list[0][i][1]+'--'+diff_list[0][i][2]+'--'+str(j+1)
			if flag not in draw_pos:
				draw_pos[flag]=[diff_list[0][i][0]]
			elif flag in draw_pos and len(draw_pos[flag])<num:
				draw_pos[flag].append(diff_list[0][i][0])

		if bool(move_list):
			for m_list in move_list:
				move_flag= m_list[1]+'--'+m_list[3]
				if move_flag not in draw_pos:
					draw_pos[move_flag]=[m_list[2]]
				else:
					draw_pos[move_flag].append(m_list[2])
		else:
			pass



		for k in range(1,3):
			for j in range(len(diff_list[k])):
				f=str(k)+diff_list[k][j][1]
				if f not in draw_pos:
					draw_pos[f] =[diff_list[k][j][0]]
				elif f in draw_pos and len(draw_pos[f])<num:
					draw_pos[f].append(diff_list[k][j][0])
					

		all_draw_pos[name]=draw_pos
	with open ('all_draw_pos.py','w',encoding='utf-8') as df:
		df.write('all_draw_pos='+pprint.pformat(all_draw_pos))

if __name__ == '__main__':
	write_r_excel('20170926')
	#get_lot_error_dict(200)
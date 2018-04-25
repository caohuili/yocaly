import openpyxl

def compare_event():
    wb=openpyxl.load_workbook('E:\\xindian\\lepu-yocaly-xml-pdf\\run_result_file\\20170921-50\\20170919-0921-yocaly-35人事件对比统计_pdf_result.xlsx')
    sheet_event = wb.get_sheet_by_name('事件')
    yocaly={}
    old={}
    new={}
    lines=sheet_event.max_row
    cols=sheet_event.max_column
    for col in range(3,cols+1):
        yocaly_l = []
        old_match = []
        old_more=[]
        old_less=[]
        new_match = []
        new_more=[]
        new_less=[]
        for line in range(3,lines+1,3):
            name = sheet_event.cell(row=line,column=2).value
            yocaly_e = sheet_event.cell(row=line,column=col).value
            new_e = sheet_event.cell(row=line+1, column=col).value
            old_e = sheet_event.cell(row=line+2, column=col).value
            if yocaly_e == 'Y':
                yocaly_l.append(name)
                if new_e =='Y':
                    new_match.append(name)
                else:
                    new_less.append(name)
                if old_e =='Y':
                    old_match.append(name)
                else:
                    old_less.append(name)
            else:
                if new_e =='Y':
                    new_more.append(name)
                else:
                    pass
                if old_e =='Y':
                    old_more.append(name)
                else:
                    pass
        yocaly[col]=yocaly_l
        old[col]=[old_match,old_more,old_less]
        new[col]=[new_match,new_more,new_less]
    print(new)
    return yocaly,new,old,cols

def write_compare_result():
    wb2 = openpyxl.Workbook()
    sheet2 = wb2.get_active_sheet()
    yocaly_d, new_d, old_d, cols = compare_event()
    line=3
    col=1
    for i in range(3,cols+1):
        sheet2.cell(row=line, column=col).value=len(yocaly_d[i])
        for n in range(len(yocaly_d[i])):
            sheet2.cell(row=line+1+n, column=col).value = yocaly_d[i][n]
        for j in range(3):
            sheet2.cell(row=line, column=col+1+j).value = len(new_d[i][j])
            for k in range(len(new_d[i][j])):
                sheet2.cell(row=line+1+k, column=col + 1 + j).value = new_d[i][j][k]
            sheet2.cell(row=line, column=col + 4 + j).value = len(old_d[i][j])
            for m in range(len(old_d[i][j])):
                sheet2.cell(row=line + 1 + m, column=col + 4 + j).value = old_d[i][j][m]
        col+=8
            #sheet2.cell(row=line, column=col+1+j).value = len(new_d[i][1])
            #sheet2.cell(row=line, column=col+1+j).value = len(new_d[i][2])
    wb2.save('E:\\xindian\\lepu-yocaly-xml-pdf\\run_result_file\\名单2.xlsx')

if __name__ == '__main__':
    write_compare_result()
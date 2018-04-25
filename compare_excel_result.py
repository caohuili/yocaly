import openpyxl

def compare_event():
    wb=openpyxl.load_workbook('E:\\pdf-xml-git\\ecg_tools\\AI_count\\201710月份统计.xlsx')
    sheet_event = wb.get_sheet_by_name('事件')
    yocaly={}
    ai={}
    lines=sheet_event.max_row
    cols=sheet_event.max_column
    for col in range(4,cols+1):
        yocaly_l = []
        ai_match = []
        ai_more=[]
        ai_less=[]
        for line in range(3,lines+1,2):
            name = sheet_event.cell(row=line,column=2).value
            yocaly_e = sheet_event.cell(row=line,column=col).value
            ai_e = sheet_event.cell(row=line+1, column=col).value
            if yocaly_e == 'Y':
                yocaly_l.append(name)
                if ai_e =='Y':
                    ai_match.append(name)
                else:
                    ai_less.append(name)
            else:
                if ai_e =='Y':
                    ai_more.append(name)
                else:
                    pass
        yocaly[col]=yocaly_l
        ai[col]=[ai_match,ai_more,ai_less]
    return yocaly,ai,cols

def write_compare_result():
    wb2 = openpyxl.Workbook()
    sheet2 = wb2.get_active_sheet()
    yocaly_d, ai_d, cols = compare_event()

    col=2
    for i in range(4,cols+1):
        line =4
        sheet2.cell(row=line-1, column=col).value=len(yocaly_d[i])
        # for n in range(len(yocaly_d[i])):
        #     sheet2.cell(row=line+1+n, column=col).value = yocaly_d[i][n]
        for j in range(3):
            sheet2.cell(row=line, column=col+j).value = len(ai_d[i][j])
            for k in range(len(ai_d[i][j])):
                sheet2.cell(row=line+1+k, column=col + j).value = ai_d[i][j][k]
                #line+=1
            #sheet2.cell(row=line, column=col + 4 + j).value = len(ai_d[i][j])
        col+=3
    wb2.save('名单5.xlsx')

if __name__ == '__main__':
    write_compare_result()